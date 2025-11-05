import openpyxl
import requests
import json
from pathlib import Path

def clean_name(name):
    return name.strip()

def get_email_lookup(mapping_sheet):
    email_lookup = {}
    for row in mapping_sheet.iter_rows(min_row=2, values_only=True):
        name, email = row[0], row[1]
        if name and email:
            email_lookup[clean_name(name)] = email.strip()
    return email_lookup

def process_account_names_sheet(sheet, email_lookup):
    data_to_post = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        account_name = row[0]
        account_lead = row[5]  # Column F
        sme_pool_raw = row[6]  # Column G

        if not account_name or not account_lead or not sme_pool_raw:
            continue

        account_lead_clean = clean_name(account_lead)
        sme_names_raw = sme_pool_raw.split(",")
        sme_names = [clean_name(name) for name in sme_names_raw if name.strip()]
        sme_combined = [account_lead_clean] + sme_names

        email_list = [email_lookup.get(name, "") for name in sme_combined if name in email_lookup]

        json_obj = {
            "Account Name": account_name.strip(),
            "SME": sme_combined,
            "Email": email_list
        }
        data_to_post.append(json_obj)
    return data_to_post

def get_unique_smes(mapping_sheet):
    sme_data = {}
    for row in mapping_sheet.iter_rows(min_row=2, values_only=True):
        name, email = row[0], row[1]
        if name and email:
            cleaned_name = clean_name(name)
            sme_data[cleaned_name] = email.strip()
    return sme_data

def ensure_db_exists(couchdb_url, db_name, username, password):
    db_url = f"{couchdb_url}/{db_name}"
    response = requests.get(db_url, auth=(username, password))
    if response.status_code == 404:
        create_response = requests.put(db_url, auth=(username, password))
        if create_response.status_code not in (200, 201):
            raise Exception(f"Failed to create database '{db_name}', Status: {create_response.status_code}, Error: {create_response.text}")
    elif response.status_code not in (200, 201):
        raise Exception(f"Error checking database: {response.status_code}, {response.text}")

def find_existing_doc(couchdb_url, db_name, username, password, key, value):
    url = f"{couchdb_url}/{db_name}/_find"
    headers = {"Content-Type": "application/json"}
    query = {
        "selector": {
            key: value
        },
        "limit": 1
    }
    response = requests.post(url, auth=(username, password), headers=headers, data=json.dumps(query))
    if response.status_code == 200:
        result = response.json()
        docs = result.get("docs", [])
        if docs:
            return docs[0]
    return None

def post_or_update_to_couchdb(couchdb_url, db_name, username, password, data_list, key_field):
    for data in data_list:
        existing_doc = find_existing_doc(couchdb_url, db_name, username, password, key_field, data[key_field])
        if existing_doc:
            data["_id"], data["_rev"] = existing_doc["_id"], existing_doc["_rev"]
            update_url = f"{couchdb_url}/{db_name}/{existing_doc['_id']}"
            response = requests.put(update_url, auth=(username, password), headers={"Content-Type": "application/json"}, data=json.dumps(data))
            if response.status_code in (200, 201):
                print(f"Updated document: {data[key_field]}")
            else:
                print(f"Failed to update: {data[key_field]}, Status: {response.status_code}, Error: {response.text}")
        else:
            post_url = f"{couchdb_url}/{db_name}"
            response = requests.post(post_url, auth=(username, password), headers={"Content-Type": "application/json"}, data=json.dumps(data))
            if response.status_code in (200, 201):
                print(f"Created new document: {data[key_field]}")
            else:
                print(f"Failed to create: {data[key_field]}, Status: {response.status_code}, Error: {response.text}")

def process_files_in_folder(folder_path, couchdb_url, username, password):
    # Ensure both databases exist
    ensure_db_exists(couchdb_url, "per_account_distribution_sheet_master", username, password)
    ensure_db_exists(couchdb_url, "per_sme_master", username, password)
    xlsx_files = list(folder_path.glob("*.xlsx"))

    if not xlsx_files:
        print("No Account Distribution Sheet found to process.")
    else:
        for file_path in xlsx_files:
            try:
                print(f"Processing file: {file_path.name}")
                wb = openpyxl.load_workbook(file_path, data_only=True)

                account_names_sheet = wb["Account Names"]
                mapping_sheet = wb["Mapping"]

                # Process SME mapping
                email_lookup = get_email_lookup(mapping_sheet)
                sme_data = get_unique_smes(mapping_sheet)
        
                sme_master_data = {
                    "Document Type": "SME Master",
                    "SMEs": sme_data
                }
                post_or_update_to_couchdb(couchdb_url, "per_sme_master", username, password, [sme_master_data], "Document Type")

                # Process account names
                data_to_post = process_account_names_sheet(account_names_sheet, email_lookup)
                post_or_update_to_couchdb(couchdb_url, "per_account_distribution_sheet_master", username, password, data_to_post, "Account Name")

                wb.close()
                file_path.unlink()
                print(f"Successfully processed and deleted: {file_path.name}\n")

            except Exception as e:
                print(f"Error processing {file_path.name}: {e}")

if __name__ == "__main__":
    couchdb_url = "http://9.20.195.22:5984"
    # db_name = "per_account_distribution_sheet_master"
    username = "admin"
    password = "admin123"
    script_dir = Path(__file__).resolve().parent
    #folder_path = "/Users/lavanyam/Downloads/Smart Triage Final"
    folder_path = script_dir / "PER Account Distribution Sheet"
    #account_names_file = "/Users/lavanyam/Downloads/Smart Triage Final/PER_Account_Distribution_Sheet_2025_V2.xlsx"
    process_files_in_folder(folder_path, couchdb_url, username, password)
    # process_files_in_folder(folder_path, couchdb_url, db_name, username, password, account_names_file)
